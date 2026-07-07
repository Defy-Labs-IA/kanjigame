# -*- coding: utf-8 -*-
"""
gerar_planilha_revisao.py
-------------------------
Extrai APENAS as linhas marcadas "Revisar leitura (auto)" para uma planilha enxuta,
pensada para um revisor que saiba japones. Cada linha ja vem com a proposta preenchida:
se estiver certa, o revisor so confirma; se errada, corrige a coluna "Leitura aprovada".

Uso:
  python gerar_planilha_revisao.py
  python gerar_planilha_revisao.py --limit 50    # so o MVP
"""
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from kanji_common import SHEET_NAME, default_workbook

SRC_DEFAULT = default_workbook().with_name(
    default_workbook().stem + "_v2_auto.xlsx")

# origem (indices 1-based no Catalogo_300)
SRC = dict(id=1, kanji=2, signif=3, hira=4, romaji=5, on=6, kun=7,
           tracos=8, mundo=11, fase=12, status=17, obs=20)

HDR = ["ID", "Kanji", "Significado PT-BR", "Mundo", "Fase",
       "Proposta (hiragana)", "Proposta (romaji)",
       "On'yomi (candidatas)", "Kun'yomi (candidatas)",
       "➜ Leitura aprovada (hiragana)", "Aprovado?", "Notas do revisor"]

INPUT_COLS = (10, 11, 12)  # colunas que o revisor preenche (destaque visual)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=str(SRC_DEFAULT))
    ap.add_argument("--out", default=None)
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()

    src_path = Path(args.src)
    out_path = Path(args.out) if args.out else src_path.with_name("Revisao_Leituras.xlsx")

    wb_in = openpyxl.load_workbook(src_path)
    ws_in = wb_in[SHEET_NAME]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revisao"

    # estilos
    head_fill = PatternFill("solid", fgColor="1F3A5F")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    input_fill = PatternFill("solid", fgColor="FFF3CD")  # amarelo suave = preencher
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="center", wrap_text=True)

    for c, title in enumerate(HDR, start=1):
        cell = ws.cell(1, c, title)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border

    out_r = 2
    for r in range(2, ws_in.max_row + 1):
        if str(ws_in.cell(r, SRC["status"]).value) != "Revisar leitura (auto)":
            continue
        if args.limit and out_r - 2 >= args.limit:
            break
        g = lambda key: ws_in.cell(r, SRC[key]).value
        proposta = g("hira")
        vals = [g("id"), g("kanji"), g("signif"), g("mundo"), g("fase"),
                proposta, g("romaji"), g("on"), g("kun"),
                proposta,          # leitura aprovada = proposta (pre-preenchida)
                "Pendente", ""]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(out_r, c, v)
            cell.border = border
            cell.alignment = center if c in (1, 2, 4, 5, 11) else wrap
            if c in INPUT_COLS:
                cell.fill = input_fill
            if c == 2:  # kanji grande
                cell.font = Font(size=18)
        out_r += 1

    # dropdown na coluna "Aprovado?"
    dv = DataValidation(type="list", formula1='"Pendente,Aprovado,Corrigido"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"K2:K{out_r - 1}")

    # larguras
    widths = [8, 8, 20, 18, 6, 16, 16, 22, 26, 20, 12, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{out_r - 1}"

    wb.save(out_path)
    print(f"Planilha de revisao gerada: {out_path}")
    print(f"Linhas para revisar: {out_r - 2}")
    print("Preencha as colunas em amarelo. Depois rode: python aplicar_revisao.py")


if __name__ == "__main__":
    main()
