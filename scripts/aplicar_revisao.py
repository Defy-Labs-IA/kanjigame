# -*- coding: utf-8 -*-
"""
aplicar_revisao.py
------------------
Le a planilha Revisao_Leituras.xlsx preenchida pelo revisor e grava as leituras
aprovadas de volta no catalogo, recalculando o romaji a partir do hiragana aprovado
e mudando o Status de "Revisar leitura (auto)" para "Validado (revisor)".

So aplica linhas cujo "Aprovado?" seja Aprovado ou Corrigido (ignora Pendente).

Uso:
  python aplicar_revisao.py
"""
import argparse
from pathlib import Path

import openpyxl

from kanji_common import COL, SHEET_NAME, default_workbook, hiragana_to_romaji

CAT_DEFAULT = default_workbook().with_name(default_workbook().stem + "_v2_auto.xlsx")
REV_DEFAULT = default_workbook().with_name("Revisao_Leituras.xlsx")

REV = dict(id=1, hira_aprov=10, status=11, notas=12)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--catalogo", default=str(CAT_DEFAULT))
    ap.add_argument("--revisao", default=str(REV_DEFAULT))
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    cat_path = Path(args.catalogo)
    out_path = Path(args.out) if args.out else cat_path.with_name(cat_path.stem + "_revisado.xlsx")

    rev = openpyxl.load_workbook(args.revisao)["Revisao"]
    aprovadas = {}
    for r in range(2, rev.max_row + 1):
        kid = rev.cell(r, REV["id"]).value
        status = str(rev.cell(r, REV["status"]).value or "").strip()
        hira = rev.cell(r, REV["hira_aprov"]).value
        if kid and hira and status in ("Aprovado", "Corrigido"):
            aprovadas[kid] = (str(hira).strip(), rev.cell(r, REV["notas"]).value)

    wb = openpyxl.load_workbook(cat_path)
    ws = wb[SHEET_NAME]
    applied = 0
    for r in range(2, ws.max_row + 1):
        kid = ws.cell(r, COL["id"]).value
        if kid in aprovadas:
            hira, notas = aprovadas[kid]
            ws.cell(r, COL["hiragana"]).value = hira
            ws.cell(r, COL["romaji"]).value = hiragana_to_romaji(hira)
            ws.cell(r, COL["status"]).value = "Validado (revisor)"
            if notas:
                ws.cell(r, COL["obs"]).value = str(notas)
            applied += 1

    wb.save(out_path)
    print(f"Leituras aprovadas aplicadas: {applied}")
    print(f"Catalogo revisado salvo em: {out_path}")
    print("Agora gere o audio definitivo: python gerar_audio.py --engine gcloud "
          f'--workbook "{out_path}"')


if __name__ == "__main__":
    main()
