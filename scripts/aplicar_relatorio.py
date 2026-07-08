# -*- coding: utf-8 -*-
"""
aplicar_relatorio.py
--------------------
Aplica as correções do Relatorio_Leituras_300.xlsx (coluna "✏ Correção") de volta:
atualiza DB (Supabase) + catalogo.json + regenera o MP3 dos kanji corrigidos.

Só processa linhas com a coluna Correção (J) preenchida.
Uso: python aplicar_relatorio.py
"""
import json
import subprocess
import sys
import urllib.request
from pathlib import Path

import openpyxl
from kanji_common import hiragana_to_romaji

RAIZ = Path(__file__).resolve().parent.parent
ENV = RAIZ / "web" / ".env.local"
REL = RAIZ / "Relatorio_Leituras_300.xlsx"
AUDIO_DIR = RAIZ / "web" / "public" / "audio"
JSONS = [RAIZ / "web" / "data" / "catalogo.json", RAIZ / "web" / "public" / "data" / "catalogo.json"]
VOZ = "ja-JP-NanamiNeural"
COL_ID, COL_CORR, COL_NOTAS = 3, 10, 12


def ler_env():
    env = {}
    for l in ENV.read_text(encoding="utf-8").splitlines():
        l = l.strip()
        if l and not l.startswith("#") and "=" in l:
            k, v = l.split("=", 1); env[k.strip()] = v.strip()
    return env


def main():
    env = ler_env()
    url = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/"); sr = env["SUPABASE_SERVICE_ROLE_KEY"]
    ws = openpyxl.load_workbook(REL)["Leituras 300"]

    correcoes = {}
    for r in range(2, ws.max_row + 1):
        kid = ws.cell(r, COL_ID).value
        corr = ws.cell(r, COL_CORR).value
        if kid and corr and str(corr).strip():
            correcoes[kid] = (str(corr).strip(), ws.cell(r, COL_NOTAS).value)

    if not correcoes:
        print("Nenhuma correção preenchida na coluna '✏ Correção'. Nada a fazer.")
        return

    # DB + JSON
    AUDIO_DIR.mkdir(parents=True, exist_ok=True)
    for kid, (hira, _notas) in correcoes.items():
        payload = {"hiragana": hira, "romaji": hiragana_to_romaji(hira), "status": "Validado (revisor)"}
        req = urllib.request.Request(f"{url}/rest/v1/kanji?id=eq.{kid}", data=json.dumps(payload).encode("utf-8"),
            method="PATCH", headers={"apikey": sr, "Authorization": f"Bearer {sr}",
                                     "Content-Type": "application/json", "Prefer": "return=minimal"})
        urllib.request.urlopen(req).read()
    for jp in JSONS:
        if not jp.exists(): continue
        dados = json.loads(jp.read_text(encoding="utf-8"))
        for k in dados:
            if k["id"] in correcoes:
                h = correcoes[k["id"]][0]
                k["hiragana"] = h; k["romaji"] = hiragana_to_romaji(h)
        jp.write_text(json.dumps(dados, ensure_ascii=False, indent=1), encoding="utf-8")

    # Áudio dos corrigidos
    for kid, (hira, _) in correcoes.items():
        subprocess.run([sys.executable, "-m", "edge_tts", "--voice", VOZ, "--text", hira,
                        "--write-media", str(AUDIO_DIR / f"{kid}.mp3")], check=True, capture_output=True)

    print(f"Correções aplicadas: {len(correcoes)} (DB + catalogo.json + MP3 regenerado).")
    print("IDs:", ", ".join(sorted(correcoes)))


if __name__ == "__main__":
    main()
