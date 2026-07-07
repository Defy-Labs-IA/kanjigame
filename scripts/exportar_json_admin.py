# -*- coding: utf-8 -*-
"""
exportar_json_admin.py
----------------------
Exporta o catalogo preenchido para admin/data/catalogo.json, consumido pelo
painel de revisao (admin/index.html).

Uso:
  python exportar_json_admin.py
"""
import argparse
import json
from pathlib import Path

import openpyxl

from kanji_common import COL, SHEET_NAME, default_workbook

SRC_DEFAULT = default_workbook().with_name(default_workbook().stem + "_v2_auto.xlsx")
OUT_DEFAULT = Path(__file__).resolve().parent.parent / "admin" / "data" / "catalogo.json"

# indices extras nao presentes em COL
IDX_MUNDO_NOME = 11
IDX_FASE = 12


def split(v):
    if not v:
        return []
    return [x.strip() for x in str(v).split(";") if x.strip()]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=str(SRC_DEFAULT))
    ap.add_argument("--out", default=str(OUT_DEFAULT))
    args = ap.parse_args()

    ws = openpyxl.load_workbook(args.src)[SHEET_NAME]
    rows = []
    for r in range(2, ws.max_row + 1):
        kid = ws.cell(r, COL["id"]).value
        kanji = ws.cell(r, COL["kanji"]).value
        if not kid or not kanji:
            continue
        status = str(ws.cell(r, COL["status"]).value or "")
        rows.append({
            "id": kid,
            "kanji": kanji,
            "significado": ws.cell(r, COL["significado"]).value or "",
            "mundo": ws.cell(r, IDX_MUNDO_NOME).value or "",
            "fase": ws.cell(r, IDX_FASE).value or "",
            "hiragana": ws.cell(r, COL["hiragana"]).value or "",
            "romaji": ws.cell(r, COL["romaji"]).value or "",
            "on": split(ws.cell(r, COL["onyomi"]).value),
            "kun": split(ws.cell(r, COL["kunyomi"]).value),
            "tracos": ws.cell(r, COL["tracos"]).value,
            "needsReview": status == "Revisar leitura (auto)",
            "status": status,
        })

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(rows, ensure_ascii=False, indent=1), encoding="utf-8")
    n_rev = sum(1 for x in rows if x["needsReview"])
    print(f"Exportado: {out}")
    print(f"Total: {len(rows)} kanjis | a revisar: {n_rev}")


if __name__ == "__main__":
    main()
