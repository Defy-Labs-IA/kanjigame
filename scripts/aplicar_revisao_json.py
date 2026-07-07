# -*- coding: utf-8 -*-
"""
aplicar_revisao_json.py
-----------------------
Aplica o arquivo decisoes.json (exportado pelo painel admin/index.html) de volta
ao catalogo, recalculando o romaji e marcando Status "Validado (revisor)".

Uso:
  python aplicar_revisao_json.py --decisoes ~/Downloads/decisoes.json
"""
import argparse
import json
from pathlib import Path

import openpyxl

from kanji_common import COL, SHEET_NAME, default_workbook, hiragana_to_romaji

CAT_DEFAULT = default_workbook().with_name(default_workbook().stem + "_v2_auto.xlsx")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--decisoes", required=True, help="caminho do decisoes.json")
    ap.add_argument("--catalogo", default=str(CAT_DEFAULT))
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    cat_path = Path(args.catalogo)
    out_path = Path(args.out) if args.out else cat_path.with_name(cat_path.stem + "_revisado.xlsx")

    decisoes = json.loads(Path(args.decisoes).read_text(encoding="utf-8"))
    by_id = {d["id"]: d for d in decisoes if d.get("hiragana")}

    wb = openpyxl.load_workbook(cat_path)
    ws = wb[SHEET_NAME]
    applied = 0
    for r in range(2, ws.max_row + 1):
        kid = ws.cell(r, COL["id"]).value
        if kid in by_id:
            d = by_id[kid]
            hira = str(d["hiragana"]).strip()
            ws.cell(r, COL["hiragana"]).value = hira
            ws.cell(r, COL["romaji"]).value = hiragana_to_romaji(hira)
            ws.cell(r, COL["status"]).value = "Validado (revisor)"
            if d.get("notas"):
                ws.cell(r, COL["obs"]).value = d["notas"]
            applied += 1

    wb.save(out_path)
    print(f"Decisoes aplicadas: {applied}")
    print(f"Catalogo revisado: {out_path}")


if __name__ == "__main__":
    main()
