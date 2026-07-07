# -*- coding: utf-8 -*-
"""
gerar_audio.py
--------------
Gera MP3 de pronuncia para cada kanji do catalogo, salvando em assets/audio/{ID}.mp3.

IMPORTANTE: o texto enviado ao TTS e a LEITURA EM HIRAGANA (coluna D), NUNCA o kanji.
Isso garante que a pronuncia corresponda a leitura ensinada (evita o TTS escolher
outra leitura do kanji).

Motores (--engine):
  edge   : edge-tts (Microsoft Edge, Neural, GRATIS, sem chave). Ideal para prototipo.
           Requer:  pip install edge-tts
  gcloud : Google Cloud Text-to-Speech (free tier generoso). Direito comercial claro.
           Requer:  variavel de ambiente GOOGLE_TTS_API_KEY
  azure  : Azure Speech (free tier). Direito comercial claro.
           Requer:  AZURE_TTS_KEY e AZURE_TTS_REGION (ex.: brazilsouth)

Uso:
  python gerar_audio.py --engine edge
  python gerar_audio.py --engine edge --limit 50
  python gerar_audio.py --engine gcloud
  python gerar_audio.py --engine azure
"""
import argparse
import base64
import json
import os
import subprocess
import sys
import urllib.request
from pathlib import Path

import openpyxl

from kanji_common import COL, SHEET_NAME, default_workbook, HEADERS

OUT_DIR = Path(__file__).resolve().parent.parent / "assets" / "audio"

DEFAULT_VOICE = {
    "edge": "ja-JP-NanamiNeural",
    "gcloud": "ja-JP-Neural2-B",
    "azure": "ja-JP-NanamiNeural",
}


# ---------------------------------------------------------------------------
# Motores
# ---------------------------------------------------------------------------
def tts_edge(text, voice, dest):
    """Chama o modulo edge_tts como subprocesso (sem chave de API)."""
    subprocess.run(
        [sys.executable, "-m", "edge_tts", "--voice", voice,
         "--text", text, "--write-media", str(dest)],
        check=True, capture_output=True,
    )


def tts_gcloud(text, voice, dest):
    key = os.environ.get("GOOGLE_TTS_API_KEY")
    if not key:
        raise RuntimeError("Defina GOOGLE_TTS_API_KEY no ambiente.")
    url = f"https://texttospeech.googleapis.com/v1/text:synthesize?key={key}"
    body = json.dumps({
        "input": {"text": text},
        "voice": {"languageCode": "ja-JP", "name": voice},
        "audioConfig": {"audioEncoding": "MP3"},
    }).encode("utf-8")
    req = urllib.request.Request(url, data=body,
                                headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as r:
        audio = json.load(r)["audioContent"]
    dest.write_bytes(base64.b64decode(audio))


def _azure_token(key, region):
    url = f"https://{region}.api.cognitive.microsoft.com/sts/v1.0/issueToken"
    req = urllib.request.Request(url, data=b"",
                                 headers={"Ocp-Apim-Subscription-Key": key})
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.read().decode("utf-8")


def tts_azure(text, voice, dest):
    key = os.environ.get("AZURE_TTS_KEY")
    region = os.environ.get("AZURE_TTS_REGION")
    if not key or not region:
        raise RuntimeError("Defina AZURE_TTS_KEY e AZURE_TTS_REGION no ambiente.")
    token = _azure_token(key, region)
    ssml = (f"<speak version='1.0' xml:lang='ja-JP'>"
            f"<voice xml:lang='ja-JP' name='{voice}'>{text}</voice></speak>")
    url = f"https://{region}.tts.speech.microsoft.com/cognitiveservices/v1"
    req = urllib.request.Request(
        url, data=ssml.encode("utf-8"),
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/ssml+xml",
            "X-Microsoft-OutputFormat": "audio-24khz-48kbitrate-mono-mp3",
            "User-Agent": HEADERS["User-Agent"],
        })
    with urllib.request.urlopen(req, timeout=30) as r:
        dest.write_bytes(r.read())


ENGINES = {"edge": tts_edge, "gcloud": tts_gcloud, "azure": tts_azure}


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--engine", choices=list(ENGINES), required=True)
    ap.add_argument("--voice", default=None)
    ap.add_argument("--workbook", default=str(default_workbook()))
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--overwrite", action="store_true")
    args = ap.parse_args()

    voice = args.voice or DEFAULT_VOICE[args.engine]
    fn = ENGINES[args.engine]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(args.workbook)
    ws = wb[SHEET_NAME]

    processed = generated = skipped = no_reading = errors = 0
    for r in range(2, ws.max_row + 1):
        kid = ws.cell(r, COL["id"]).value
        hira = ws.cell(r, COL["hiragana"]).value
        if not kid:
            continue
        if args.limit and processed >= args.limit:
            break
        processed += 1

        if not hira:
            no_reading += 1
            print(f"  [SEM LEITURA] {kid}: coluna hiragana vazia -> rode preencher_catalogo antes")
            continue

        dest = OUT_DIR / f"{kid}.mp3"
        if dest.exists() and not args.overwrite:
            skipped += 1
            continue

        try:
            fn(str(hira), voice, dest)
            generated += 1
        except Exception as e:  # noqa: BLE001
            errors += 1
            print(f"  [ERRO] {kid} ({hira}): {e}")

    print(f"\n=== gerar_audio ({args.engine}, voz {voice}) ===")
    print(f"Processados   : {processed}")
    print(f"Gerados       : {generated}")
    print(f"Ja existiam   : {skipped}")
    print(f"Sem leitura   : {no_reading}")
    print(f"Erros         : {errors}")
    print(f"Pasta         : {OUT_DIR}")


if __name__ == "__main__":
    main()
