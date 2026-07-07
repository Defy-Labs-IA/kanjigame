# -*- coding: utf-8 -*-
"""
preencher_catalogo.py
---------------------
Le os kanjis da planilha e preenche, via kanjiapi.dev (dados KANJIDIC2 / EDRDG, CC BY-SA 4.0):
  - On'yomi, Kun'yomi, N de tracos  -> FATOS (preenchidos direto)
  - Leitura principal (hiragana) + Romaji -> PROPOSTA automatica (kun primeiro, senao on)
    -> toda linha auto-preenchida recebe Status "Revisar leitura (auto)" e as
       leituras candidatas ficam em Observacoes, para revisao humana/nativa.

NAO sobrescreve celulas ja preenchidas por humanos (a menos que --overwrite).
Salva num arquivo NOVO (nao destroi o original). Respostas da API sao cacheadas.

Uso:
  python preencher_catalogo.py                 # todos os kanjis
  python preencher_catalogo.py --limit 50      # so os 50 primeiros (MVP)
  python preencher_catalogo.py --overwrite     # regrava tudo
"""
import argparse
import json
from pathlib import Path

import openpyxl

from kanji_common import (
    COL, SHEET_NAME, default_workbook, http_get_json,
    katakana_to_hiragana, hiragana_to_romaji, clean_kun,
)

CACHE_DIR = Path(__file__).resolve().parent / "cache" / "kanjiapi"


def fetch_kanji(kanji):
    """Busca (com cache local) o registro do kanji na kanjiapi.dev."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cp = f"{ord(kanji):x}"
    cache_file = CACHE_DIR / f"{cp}.json"
    if cache_file.exists():
        return json.loads(cache_file.read_text(encoding="utf-8"))
    import urllib.parse
    url = "https://kanjiapi.dev/v1/kanji/" + urllib.parse.quote(kanji)
    data = http_get_json(url)
    cache_file.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return data


def _pick(readings):
    """Escolhe a melhor leitura 'solta' de uma lista KANJIDIC.
    Prefere formas sem marcador de afixo (- no inicio/fim); remove okurigana (apos '.')."""
    if not readings:
        return ""
    standalone = [x for x in readings if not (x.startswith("-") or x.endswith("-"))]
    pool = standalone or readings
    return clean_kun(pool[0])  # clean_kun tira '.', '-' e espacos


def propose_reading(on_readings, kun_readings):
    """Retorna (hiragana, romaji) proposto. Kun tem prioridade; senao on (convertido)."""
    hira = _pick(kun_readings)
    if not hira and on_readings:
        hira = _pick([katakana_to_hiragana(x) for x in on_readings])
    if not hira:
        return "", ""
    return hira, hiragana_to_romaji(hira)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=str(default_workbook()))
    ap.add_argument("--out", default=None, help="arquivo de saida (default: *_v2_auto.xlsx)")
    ap.add_argument("--limit", type=int, default=None, help="processar apenas N linhas")
    ap.add_argument("--overwrite", action="store_true", help="sobrescrever celulas ja preenchidas")
    args = ap.parse_args()

    wb_path = Path(args.workbook)
    out_path = Path(args.out) if args.out else wb_path.with_name(wb_path.stem + "_v2_auto.xlsx")

    wb = openpyxl.load_workbook(wb_path)
    ws = wb[SHEET_NAME]

    processed = filled_facts = proposed = skipped = errors = 0
    for r in range(2, ws.max_row + 1):
        kanji = ws.cell(r, COL["kanji"]).value
        if not kanji:
            continue
        if args.limit and processed >= args.limit:
            break
        processed += 1

        already = ws.cell(r, COL["onyomi"]).value and ws.cell(r, COL["hiragana"]).value
        if already and not args.overwrite:
            skipped += 1
            continue

        try:
            d = fetch_kanji(kanji)
        except Exception as e:  # noqa: BLE001
            errors += 1
            print(f"  [ERRO] {kanji} (linha {r}): {e}")
            continue

        on = d.get("on_readings") or []
        kun = d.get("kun_readings") or []
        strokes = d.get("stroke_count")

        # Fatos
        ws.cell(r, COL["onyomi"]).value = "; ".join(on)
        ws.cell(r, COL["kunyomi"]).value = "; ".join(kun)
        if strokes is not None:
            ws.cell(r, COL["tracos"]).value = strokes
        filled_facts += 1

        # Proposta de leitura principal (so se estiver vazia, salvo overwrite)
        if args.overwrite or not ws.cell(r, COL["hiragana"]).value:
            hira, romaji = propose_reading(on, kun)
            if hira:
                ws.cell(r, COL["hiragana"]).value = hira
                ws.cell(r, COL["romaji"]).value = romaji
                ws.cell(r, COL["status"]).value = "Revisar leitura (auto)"
                cand = f"on: {', '.join(on) or '-'} | kun: {', '.join(kun) or '-'}"
                ws.cell(r, COL["obs"]).value = f"Proposta automatica. Candidatas -> {cand}"
                proposed += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print("\n=== preencher_catalogo ===")
    print(f"Kanjis processados : {processed}")
    print(f"Fatos preenchidos  : {filled_facts} (on/kun/tracos)")
    print(f"Leituras propostas : {proposed} (marcadas 'Revisar leitura (auto)')")
    print(f"Ja preenchidos     : {skipped} (preservados)")
    print(f"Erros              : {errors}")
    print(f"Salvo em           : {out_path}")


if __name__ == "__main__":
    main()
