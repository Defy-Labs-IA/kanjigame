# -*- coding: utf-8 -*-
"""
baixar_tracos.py
----------------
Baixa do animCJK (SVGs de tracos ja prontos para animar por CSS) apenas os kanjis
do catalogo, salvando como assets/strokes/{ID}.svg (ex.: K001.svg).

Fonte : https://github.com/parsimonhi/animCJK  (pasta svgsJa)
        arquivo = unicode decimal do caractere + .svg  (ex.: 中 -> 20013.svg)
Licenca: SVGs de kanji sob Arphic Public License (exige atribuicao).

Pula arquivos ja baixados. Uso:
  python baixar_tracos.py                 # todos
  python baixar_tracos.py --limit 50      # so o MVP
"""
import argparse
import time
from pathlib import Path

import openpyxl

from kanji_common import COL, SHEET_NAME, default_workbook, http_get

BASE = "https://raw.githubusercontent.com/parsimonhi/animCJK/master/svgsJa/{dec}.svg"
OUT_DIR = Path(__file__).resolve().parent.parent / "assets" / "strokes"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=str(default_workbook()))
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--sleep", type=float, default=0.3, help="pausa entre downloads (s)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    ws = wb[SHEET_NAME]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    processed = downloaded = skipped = missing = 0
    for r in range(2, ws.max_row + 1):
        kanji = ws.cell(r, COL["kanji"]).value
        kid = ws.cell(r, COL["id"]).value
        if not kanji or not kid:
            continue
        if args.limit and processed >= args.limit:
            break
        processed += 1

        dest = OUT_DIR / f"{kid}.svg"
        if dest.exists():
            skipped += 1
            continue

        dec = ord(kanji)
        url = BASE.format(dec=dec)
        try:
            data = http_get(url)
            dest.write_bytes(data)
            downloaded += 1
        except Exception as e:  # noqa: BLE001
            missing += 1
            print(f"  [SEM SVG] {kid} {kanji} (U+{dec:X}): {e}")
        time.sleep(args.sleep)

    print("\n=== baixar_tracos ===")
    print(f"Processados : {processed}")
    print(f"Baixados    : {downloaded}")
    print(f"Ja existiam : {skipped}")
    print(f"Nao achados : {missing}")
    print(f"Pasta       : {OUT_DIR}")


if __name__ == "__main__":
    main()
