# -*- coding: utf-8 -*-
"""
gerar_relatorio_leituras.py
---------------------------
Gera Relatorio_Leituras_300.xlsx a partir do Supabase (leituras atuais) para
conferência por um nativo: kanji, leitura, romaji, significado + candidatas
on/kun e colunas para marcar correção.
"""
import json
import urllib.request
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent.parent
ENV = RAIZ / "web" / ".env.local"
OUT = RAIZ / "Relatorio_Leituras_300.xlsx"

HDR = ["Mundo", "Fase", "ID", "Kanji", "Significado", "Leitura (hiragana)", "Romaji",
       "On'yomi (candidatas)", "Kun'yomi (candidatas)",
       "✏ Correção (hiragana)", "OK?", "Notas do revisor"]
INPUT_COLS = (10, 11, 12)


def ler_env():
    env = {}
    for l in ENV.read_text(encoding="utf-8").splitlines():
        l = l.strip()
        if l and not l.startswith("#") and "=" in l:
            k, v = l.split("=", 1); env[k.strip()] = v.strip()
    return env


def buscar():
    env = ler_env()
    url = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
    sr = env["SUPABASE_SERVICE_ROLE_KEY"]
    req = urllib.request.Request(
        url + "/rest/v1/kanji?select=id,kanji,significado,mundo,fase,hiragana,romaji,onyomi,kunyomi&order=id",
        headers={"apikey": sr, "Authorization": f"Bearer {sr}"})
    return json.load(urllib.request.urlopen(req))


def main():
    dados = buscar()
    # ordena por fase depois id
    dados.sort(key=lambda k: (int(k["fase"]) if str(k["fase"]).isdigit() else 999, k["id"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leituras 300"

    head_fill = PatternFill("solid", fgColor="1F3A5F")
    head_font = Font(bold=True, color="FFFFFF")
    input_fill = PatternFill("solid", fgColor="FFF3CD")
    thin = Side(style="thin", color="D8D8D8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="center", wrap_text=True)

    for c, t in enumerate(HDR, 1):
        cell = ws.cell(1, c, t)
        cell.fill = head_fill; cell.font = head_font; cell.alignment = center; cell.border = border

    r = 2
    for k in dados:
        vals = [k.get("mundo", ""), k.get("fase", ""), k["id"], k["kanji"], k.get("significado", ""),
                k.get("hiragana", ""), k.get("romaji", ""),
                "; ".join(k.get("onyomi") or []), "; ".join(k.get("kunyomi") or []),
                "", "", ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = border
            cell.alignment = center if c in (2, 3, 4, 7, 11) else wrap
            if c in INPUT_COLS:
                cell.fill = input_fill
            if c == 4:
                cell.font = Font(size=18)
        r += 1

    dv = DataValidation(type="list", formula1='"OK,Corrigir,Dúvida"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"K2:K{r-1}")

    widths = [20, 6, 7, 8, 22, 18, 14, 20, 24, 20, 10, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{r-1}"

    wb.save(OUT)
    print(f"Relatório gerado: {OUT}  ({len(dados)} kanji)")


if __name__ == "__main__":
    main()
